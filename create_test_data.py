import os
from openpyxl import Workbook

def create_clean_data(directory):
    wb = Workbook()
    ws = wb.active
    ws.title = "Clean Data"
    
    # Clean headers on row 1
    ws.append(["Date", "Coal Consumption", "Steam Generation", "Power Generation"])
    ws.append(["2023-10-01", 1200.50, 45.2, 350.5])
    ws.append(["2023-10-02", 1190.00, 44.8, 345.0])
    
    wb.save(os.path.join(directory, "clean_data.xlsx"))

def create_messy_data(directory):
    wb = Workbook()
    ws = wb.active
    ws.title = "Messy Data"
    
    # Row 1: Title
    ws.append(["Monthly Operations Report - ACME Corp"])
    # Row 2: Blank
    ws.append([])
    # Row 3: Messy Headers
    ws.append(["Date", "Coal Used (MT)", "Steam (T/hr)", "Pwr Gen", "Comments"])
    # Row 4: Mixed data formats
    ws.append(["2023-10-01", "1,234.56", "45%", "YES", "Normal operations"])
    ws.append(["2023-10-02", "1,190.00", "N/A", "NO", "Sensor offline"])
    
    wb.save(os.path.join(directory, "messy_data.xlsx"))

def create_multi_asset_data(directory):
    wb = Workbook()
    ws = wb.active
    ws.title = "Multi Asset"
    
    # Headers with embedded assets
    ws.append(["Date", "Coal Consumption AFBC-1", "Steam (Boiler 1)", "Coal Consumption AFBC-2", "Power TG-1"])
    ws.append(["2023-10-01", 500.0, 20.0, 600.0, 150.0])
    ws.append(["2023-10-02", 490.0, 19.5, 610.0, 148.0])
    
    wb.save(os.path.join(directory, "multi_asset.xlsx"))

if __name__ == "__main__":
    # Create the test_files directory if it doesn't exist
    output_dir = "test_files"
    os.makedirs(output_dir, exist_ok=True)
    
    create_clean_data(output_dir)
    create_messy_data(output_dir)
    create_multi_asset_data(output_dir)
    print(f"Test files successfully generated in the '{output_dir}' directory.")